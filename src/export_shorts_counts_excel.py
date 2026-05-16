import argparse
import glob
import json
import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from shorts_config import CATEGORIES, COUNTRIES, OUTPUT_DIR

SKIP_FILES = {"shorts_product.json"}


def list_shorts_files():
    pattern = os.path.join(OUTPUT_DIR, "shorts_*.json")
    result = {}
    for file_path in glob.glob(pattern):
        file_name = os.path.basename(file_path)
        if file_name in SKIP_FILES:
            continue
        if not file_name.startswith("shorts_") or not file_name.endswith(".json"):
            continue
        country_id = file_name[len("shorts_") : -len(".json")]
        if not country_id:
            continue
        result[country_id] = file_path
    return result


def count_valid_entries(items):
    if not isinstance(items, list):
        return 0

    count = 0
    for item in items:
        if isinstance(item, str) and item.strip():
            count += 1
            continue
        if isinstance(item, dict):
            video_id = item.get("video_id") or item.get("id")
            if isinstance(video_id, str) and video_id.strip():
                count += 1
    return count


def load_country_category_counts(file_path):
    counts = {category: 0 for category in CATEGORIES}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        return counts

    if not isinstance(data, dict):
        return counts

    for category in CATEGORIES:
        counts[category] = count_valid_entries(data.get(category, []))
    return counts


def sort_country_ids(country_ids):
    preferred = [country["id"] for country in COUNTRIES] + ["unknown", "other"]
    order = {country_id: idx for idx, country_id in enumerate(preferred)}
    fallback_base = len(order)
    return sorted(country_ids, key=lambda cid: (order.get(cid, fallback_base), cid))


def write_excel(output_file, country_counts):
    wb = Workbook()
    ws = wb.active
    ws.title = "counts"

    categories = list(CATEGORIES.keys())
    headers = ["country"] + categories + ["total"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill

    for country_id in sort_country_ids(country_counts.keys()):
        category_counts = country_counts[country_id]
        row = [country_id]
        total = 0
        for category in categories:
            value = int(category_counts.get(category, 0))
            row.append(value)
            total += value
        row.append(total)
        ws.append(row)

    totals_row = ["ALL"]
    for category in categories:
        totals_row.append(sum(country_counts[cid].get(category, 0) for cid in country_counts))
    totals_row.append(sum(totals_row[1:]))
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="FCE4D6")

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 2, 40))

    meta = wb.create_sheet("meta")
    meta.append(["generated_at", datetime.now().isoformat(timespec="seconds")])
    meta.append(["source_pattern", os.path.join(OUTPUT_DIR, "shorts_*.json")])
    meta.append(["skipped_files", ", ".join(sorted(SKIP_FILES))])

    wb.save(output_file)


def main():
    parser = argparse.ArgumentParser(
        description="Export per-country per-category shorts counts to an Excel file."
    )
    parser.add_argument(
        "--output",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "shorts_country_category_counts.xlsx"),
        help="Output Excel file path.",
    )
    args = parser.parse_args()

    shorts_files = list_shorts_files()
    if not shorts_files:
        print("No shorts_*.json files found.")
        return

    country_counts = {
        country_id: load_country_category_counts(file_path)
        for country_id, file_path in shorts_files.items()
    }

    write_excel(args.output, country_counts)
    print(f"Exported: {args.output}")
    print(f"Countries: {len(country_counts)}")
    print(f"Categories: {len(CATEGORIES)}")


if __name__ == "__main__":
    main()